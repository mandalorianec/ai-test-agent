import os
import json
from datetime import datetime
from pprint import pprint

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

from main import generate_tests
from gemini import enhance_test_cases_with_ai


TESTIT_TEMPLATE_COLUMNS = [
    "ID",
    "Расположение",
    "Наименование",
    "Автоматизирован",
    "Предусловия",
    "Шаги",
    "Постусловия",
    "Ожидаемый результат",
    "Тестовые данные",
    "Комментарии",
    "Итерации",
    "Приоритет",
    "Статус",
    "Дата создания",
    "Автор",
    "Длительность",
    "Тег",
]


def _empty_testit_row() -> dict:
    return {column: "" for column in TESTIT_TEMPLATE_COLUMNS}


def _safe_list(value):
    return value if isinstance(value, list) else []


def _map_priority_to_ru(priority: str) -> str:
    mapping = {
        "highest": "Критический",
        "high": "Высокий",
        "medium": "Средний",
        "low": "Низкий",
        "lowest": "Незначительный",
    }
    if not priority:
        return "Средний"
    return mapping.get(str(priority).strip().lower(), str(priority))


def _build_location(test: dict) -> str:
    endpoint = test.get("endpoint", {}) or {}
    method = endpoint.get("method", "")
    path = endpoint.get("path", "")

    tags = _safe_list(test.get("tags"))
    technical_tags = {"api", "positive", "negative"}
    if method:
        technical_tags.add(method.lower())

    business_tags = [tag for tag in tags if tag not in technical_tags]

    if business_tags:
        return " -> ".join(business_tags)

    if method and path:
        return f"{method} {path}"

    return "API"


def _build_test_data(test: dict) -> str:
    metadata = test.get("metadata", {}) or {}
    parts = []

    path_params = _safe_list(metadata.get("path_params"))
    if path_params:
        parts.append("Path: " + "; ".join(path_params))

    query_params = _safe_list(metadata.get("query_params"))
    if query_params:
        parts.append("Query: " + "; ".join(query_params))

    required_body_fields = _safe_list(metadata.get("required_body_fields"))
    if required_body_fields:
        parts.append("Body: " + "; ".join(required_body_fields))

    error_code = metadata.get("error_code")
    if error_code:
        parts.append(f"Ожидаемый код ошибки: {error_code}")

    return "\n".join(parts)


def _build_iteration_value(test: dict) -> str:
    metadata = test.get("metadata", {}) or {}

    path_params = _safe_list(metadata.get("path_params"))
    query_params = _safe_list(metadata.get("query_params"))

    variables = []
    variables.extend([f"%{name}" for name in path_params if name])
    variables.extend([f"%{name}" for name in query_params if name])

    return "; ".join(variables)


def _create_test_header_row(
    test: dict,
    author: str = "",
    status: str = "Готов",
    automated: str = "Нет",
    duration: str = "",
    test_id: str = "",
) -> dict:
    row = _empty_testit_row()

    tags = _safe_list(test.get("tags"))
    endpoint = test.get("endpoint", {}) or {}

    row["ID"] = test_id
    row["Расположение"] = _build_location(test)
    row["Наименование"] = test.get("name", "")
    row["Автоматизирован"] = automated
    row["Тестовые данные"] = _build_test_data(test)
    row["Комментарии"] = test.get("description", "") or ""
    row["Итерации"] = _build_iteration_value(test)
    row["Приоритет"] = _map_priority_to_ru(test.get("priority", "medium"))
    row["Статус"] = status
    row["Дата создания"] = datetime.now().strftime("%m/%d/%Y %H:%M:%S")
    row["Автор"] = author
    row["Длительность"] = duration
    row["Тег"] = "; ".join(tags)

    if endpoint.get("method") and endpoint.get("path"):
        endpoint_text = f"{endpoint['method']} {endpoint['path']}"
        row["Комментарии"] = f"{endpoint_text}\n{row['Комментарии']}".strip()

    return row


def build_excel_table_dict(
    tests: list[dict],
    project_name: str = "Demo",
    author: str = "",
    status: str = "Готов",
    automated: str = "Нет",
    duration: str = "",
    start_id: int | None = None,
) -> dict:
    rows = []
    current_id = start_id

    for test in tests:
        test_id = ""
        if current_id is not None:
            test_id = str(current_id)
            current_id += 1

        rows.append(
            _create_test_header_row(
                test=test,
                author=author,
                status=status,
                automated=automated,
                duration=duration,
                test_id=test_id,
            )
        )

        for precondition in _safe_list(test.get("preconditions")):
            row = _empty_testit_row()
            row["Предусловия"] = str(precondition)
            row["Ожидаемый результат"] = "Предусловие выполнено"
            rows.append(row)

        for step in _safe_list(test.get("steps")):
            row = _empty_testit_row()
            row["Шаги"] = str(step.get("action", ""))
            row["Ожидаемый результат"] = str(step.get("expected_result", ""))
            rows.append(row)

        for postcondition in _safe_list(test.get("postconditions")):
            row = _empty_testit_row()
            row["Постусловия"] = str(postcondition)
            rows.append(row)

    return {
        "sheet_name": f"Project_{project_name}",
        "columns": TESTIT_TEMPLATE_COLUMNS,
        "rows": rows,
    }


def generate_excel_table_dict_from_swagger(
    swagger: dict,
    project_name: str = "Demo",
    author: str = "",
    status: str = "Готов",
    automated: str = "Нет",
    duration: str = "",
    start_id: int | None = None,
) -> dict:
    tests = generate_tests(swagger)
    enhanced_tests = enhance_test_cases_with_ai(tests)
    return build_excel_table_dict(
        tests=enhanced_tests,
        project_name=project_name,
        author=author,
        status=status,
        automated=automated,
        duration=duration,
        start_id=start_id,
    )


##########################################
"""
def preview_excel_table_dict(excel_table: dict, limit: int = 10):
    print("\n=== EXCEL TABLE PREVIEW ===")
    print(f"sheet_name: {excel_table['sheet_name']}")
    print(f"columns: {excel_table['columns']}")
    print(f"rows total: {len(excel_table['rows'])}")

    for i, row in enumerate(excel_table["rows"][:limit], start=1):
        print(f"\n--- Row {i} ---")
        pprint(row)

    print("\n===========================\n")


def create_excel_file(excel_table: dict, output_path: str = "testit_export.xlsx") -> str:

    #Создаёт xlsx-файл из словаря excel_table и возвращает путь к файлу.

    wb = Workbook()
    ws = wb.active

    sheet_name = excel_table.get("sheet_name", "Sheet1")
    ws.title = sheet_name[:31]

    columns = excel_table.get("columns", [])
    rows = excel_table.get("rows", [])

    # Header
    for col_idx, column_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=column_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")

    # Data
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, column_name in enumerate(columns, start=1):
            value = row_data.get(column_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Widths — близко к шаблону
    column_widths = {
        "A": 11,
        "B": 51,
        "C": 25,
        "D": 20,
        "E": 51,
        "F": 25,
        "G": 33,
        "H": 51,
        "I": 20,
        "J": 51,
        "K": 37,
        "L": 14,
        "M": 12,
        "N": 19,
        "O": 14,
        "P": 16,
        "Q": 14,
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Borders
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = ws.max_row
    max_col = ws.max_column

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border

    # Row heights
    ws.row_dimensions[1].height = 24
    for row_idx in range(2, max_row + 1):
        ws.row_dimensions[row_idx].height = 36

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto filter
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    # Save
    wb.save(output_path)
    return output_path

if __name__ == "__main__":
    SWAGGER_PATH = "discord-api.json"
    OUTPUT_XLSX_PATH = "testit_export.xlsx"

    if not os.path.exists(SWAGGER_PATH):
        print(f"Файл не найден: {SWAGGER_PATH}")
        raise SystemExit(1)

    with open(SWAGGER_PATH, "r", encoding="utf-8") as f:
        swagger = json.load(f)

    excel_table = generate_excel_table_dict_from_swagger(
        swagger=swagger,
        project_name="Демо-проект🚀",
        author="Артемизий",
        status="Готов",
        automated="Нет",
        duration="0h 5m 0s",
        start_id=1,
    )

    preview_excel_table_dict(excel_table, limit=12)

    created_file = create_excel_file(excel_table, OUTPUT_XLSX_PATH)
    print(f"Excel файл создан: {created_file}")
    """