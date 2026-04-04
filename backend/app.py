import asyncio
import io
import json
from urllib.parse import quote

import httpx
from fastapi import FastAPI, UploadFile, File, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

from main import generate_tests
from xlsx_maker import generate_excel_table_dict_from_swagger, enhance_test_cases_with_ai


app = FastAPI(title="AI Test Agent Backend")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://213.165.52.36",
        "http://localhost",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

TEST_IT_URL_ITEMS = "https://team-a0ff.testit.software/api/v2/workItems"
TEST_IT_API_TOKEN = "UDlWRVhxa0hrYlFqekVtWWN4"
PROJECT_ID = "019ce25b-28d8-7ff0-b14f-0d2f51c7a953"
SECTION_ID = "0ab7fd39-1730-484a-a08a-f45940fc2184"

HEADERS = {
    "Authorization": f"PrivateToken {TEST_IT_API_TOKEN}",
    "Content-Type": "application/json",
}

MAX_CONCURRENT_REQUESTS = 5
RETRIES = 3
TIMEOUT = 30.0

semaphore = asyncio.Semaphore(MAX_CONCURRENT_REQUESTS)


def validate_json_upload(file: UploadFile) -> None:
    if not file.filename:
        raise HTTPException(status_code=400, detail="Файл не выбран")

    if not file.filename.lower().endswith(".json"):
        raise HTTPException(status_code=400, detail="Нужен файл .json")


async def read_swagger_json(file: UploadFile) -> dict:
    validate_json_upload(file)

    try:
        content = await file.read()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Не удалось прочитать файл: {e}")

    try:
        return json.loads(content)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Файл содержит невалидный JSON")


def map_to_testit_format(test: dict) -> dict:
    return {
        "projectId": PROJECT_ID,
        "sectionId": SECTION_ID,
        "name": test["name"],
        "description": f"{test['endpoint']['method']} {test['endpoint']['path']}",
        "entityTypeName": "TestCases",
        "duration": 1000,
        "state": "NeedsWork",
        "priority": test.get("priority", "Medium"),
        "attributes": {},
        "tags": [{"name": tag} for tag in test.get("tags", [])],
        "preconditionSteps": [{"action": p} for p in test.get("preconditions", [])],
        "postconditionSteps": [{"action": p} for p in test.get("postconditions", [])],
        "links": [],
        "steps": [
            {
                "action": step["action"],
                "expected": step["expected_result"],
            }
            for step in test.get("steps", [])
        ],
    }


async def send_to_testit(client: httpx.AsyncClient, test: dict):
    for attempt in range(RETRIES):
        try:
            async with semaphore:
                response = await client.post(
                    TEST_IT_URL_ITEMS,
                    json=test,
                    headers=HEADERS,
                )
                return {
                    "status": response.status_code,
                    "name": test["name"],
                    "response": response.text,
                }
        except httpx.PoolTimeout:
            if attempt < RETRIES - 1:
                await asyncio.sleep(attempt + 1)
                continue
            return {
                "status": "error",
                "name": test["name"],
                "response": "PoolTimeout",
            }
        except Exception as e:
            if attempt < RETRIES - 1:
                await asyncio.sleep(attempt + 1)
                continue
            return {
                "status": "error",
                "name": test["name"],
                "response": str(e),
            }


async def send_all_tests(testit_tests: list):
    limits = httpx.Limits(
        max_connections=10,
        max_keepalive_connections=5,
    )
    async with httpx.AsyncClient(timeout=TIMEOUT, limits=limits) as client:
        tasks = [send_to_testit(client, test) for test in testit_tests]
        return await asyncio.gather(*tasks)


def create_excel_stream(excel_table: dict) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active

    sheet_name = excel_table.get("sheet_name", "Sheet1")
    ws.title = str(sheet_name)[:31]

    columns = excel_table.get("columns", [])
    rows = excel_table.get("rows", [])

    if not columns:
        raise ValueError("В excel_table отсутствуют columns")

    for col_idx, column_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=column_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, column_name in enumerate(columns, start=1):
            value = row_data.get(column_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    column_widths = {
        "A": 11,
        "B": 51,
        "C": 25,
        "D": 20,
        "E": 51,
        "F": 25,
        "G": 33,
        "H": 51,
        "I": 20,
        "J": 51,
        "K": 37,
        "L": 14,
        "M": 12,
        "N": 19,
        "O": 14,
        "P": 16,
        "Q": 14,
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    max_col = len(columns)
    for idx in range(1, max_col + 1):
        col_letter = get_column_letter(idx)
        if col_letter not in column_widths:
            ws.column_dimensions[col_letter].width = 20

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = ws.max_row
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border

    ws.row_dimensions[1].height = 24
    for row_idx in range(2, max_row + 1):
        ws.row_dimensions[row_idx].height = 36

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


@app.post("/generate")
async def generate(file: UploadFile = File(...)):
    swagger_data = await read_swagger_json(file)

    try:
        tests = generate_tests(swagger_data)
        enhanced_test = enhance_test_cases_with_ai(tests)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка генерации тестов: {e}")

    try:
        testit_tests = [map_to_testit_format(t) for t in enhanced_test]
        results = await send_all_tests(testit_tests)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка отправки в Test IT: {e}")

    created_count = sum(1 for result in results if result.get("status") in (200, 201))
    failed = [result for result in results if result.get("status") not in (200, 201)]

    return {
        "generated": len(enhanced_test),
        "created": created_count,
        "failed": failed,
    }


@app.post("/xlsx")
async def generate_xlsx(
    file: UploadFile = File(...),
    project_name: str = Form("Demo"),
    author: str = Form(""),
    status: str = Form("Готов"),
    automated: str = Form("Нет"),
    duration: str = Form(""),
    start_id: str = Form(""),
):
    swagger_data = await read_swagger_json(file)

    parsed_start_id = None
    if str(start_id).strip():
        try:
            parsed_start_id = int(start_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="start_id должен быть числом")

    try:
        excel_table = generate_excel_table_dict_from_swagger(
            swagger=swagger_data,
            project_name=project_name,
            author=author,
            status=status,
            automated=automated,
            duration=duration,
            start_id=parsed_start_id,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка генерации excel-структуры: {e}")

    try:
        xlsx_stream = create_excel_stream(excel_table)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка сборки XLSX-файла: {e}")

    output_filename = f"{project_name or 'testit_export'}.xlsx"
    encoded_filename = quote(output_filename)

    headers = {
        "Content-Disposition": (
            f'attachment; filename="{output_filename}"; '
            f"filename*=UTF-8''{encoded_filename}"
        )
    }

    return StreamingResponse(
        xlsx_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )